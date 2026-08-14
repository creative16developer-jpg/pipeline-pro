"""
Map step router — /api/pipelines/{id}/map-data  +  /api/pipelines/{id}/map-confirm
                  /api/stores/{id}/category-mappings

The Map step sits inside the existing Review pause.  The client confirms category
mappings via map-confirm, which also triggers the pipeline resume (Upload → Sync).

Multi-category support: each Sunsky category maps to a list of WooCommerce categories
with one designated as primary.  The full set is stored in woo_cats_json (JSON).
Backward-compat columns woo_cat_id / woo_cat_name mirror the primary category.
"""
from __future__ import annotations

import asyncio
import json
from datetime import datetime, timezone
from typing import Optional

import csv
import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from database import get_db
from models.models import (
    PipelineJob, Product, SunskyCategoryMapping, WooCategory, AttributeProfile, Job, JobType
)

router = APIRouter(tags=["map-step"])


# ─────────────────────────────────────────────────────────────────────────────
# Schemas
# ─────────────────────────────────────────────────────────────────────────────

class WooCatEntry(BaseModel):
    id: int
    name: str


class MappingEntry(BaseModel):
    sunsky_cat: str
    woo_cats: list[WooCatEntry] = []
    primary_woo_cat_id: Optional[int] = None
    profile_id: Optional[int] = None
    save_as_rule: bool = True


class MapConfirmRequest(BaseModel):
    mappings: list[MappingEntry] = []


class CategoryMappingUpdate(BaseModel):
    sunsky_cat: str
    woo_cats: list[WooCatEntry] = []
    primary_woo_cat_id: Optional[int] = None
    profile_id: Optional[int] = None


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _extract_sunsky_cat(raw: dict, name_map: Optional[dict[str, str]] = None) -> str:
    """Best-effort extraction of Sunsky category NAME from product raw_data.

    Real Sunsky product responses only include a numeric categoryId, not a
    name field — confirmed against live data (2026-08-01). When no name
    field is present, resolves the ID through `name_map` (built from
    sunsky_client.get_category_name_map()) if provided. This must produce
    the exact same value that gets saved as SunskyCategoryMapping.sunsky_cat
    below, or future pipeline runs won't recognize this category as already
    mapped — see services/enrich_service.py's extract_sunsky_category() for
    the sibling copy of this same fix.
    """
    for key in ("catName", "categoryName", "category_name", "cat_name"):
        v = str(raw.get(key) or "").strip()
        if v:
            return v
    cat_id = str(raw.get("categoryId") or raw.get("catId") or raw.get("category_id") or "").strip()
    if cat_id and name_map:
        name = name_map.get(cat_id)
        if name:
            return name
    return cat_id


def _extract_sunsky_cat_id(raw: dict) -> str:
    """Raw numeric Sunsky category ID from a product's raw_data — the stable
    match key sunsky_category_mappings.sunsky_cat_id exists for (see model
    docstring). Mirrors job_tasks.py's _get_sunsky_cat_id so Category Review,
    Content Review, and Sync all agree on the same value for the same product.
    """
    return str(raw.get("categoryId") or raw.get("catId") or raw.get("category_id") or "").strip()


def _mapping_woo_cats(m: SunskyCategoryMapping) -> list[dict]:
    """Return list of {id, name} dicts from a saved mapping row."""
    if m.woo_cats_json:
        try:
            return json.loads(m.woo_cats_json)
        except Exception:
            pass
    if m.woo_cat_id:
        return [{"id": m.woo_cat_id, "name": m.woo_cat_name or ""}]
    return []


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline-scoped endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/pipelines/{pipeline_id}/map-data")
async def get_map_data(pipeline_id: int, db: AsyncSession = Depends(get_db)):
    """
    Returns unique Sunsky categories found in this pipeline's product batch,
    merged with any saved mappings for this store.

    woo_options includes parent_id so the frontend can render a hierarchy tree.
    Each category entry has:
      - woo_cats: list of {id, name} (all assigned WooCommerce cats)
      - primary_woo_cat_id: which one is primary
      - is_new: True when no saved rule exists (needs manual assignment)
    """
    pl = await db.get(PipelineJob, pipeline_id)
    if not pl:
        raise HTTPException(status_code=404, detail="Pipeline not found")

    # Load products for this pipeline's fetch job
    products = (
        await db.execute(
            select(Product).where(Product.fetch_job_id == pl.fetch_job_id)
        )
    ).scalars().all()

    from services.enrich_service import get_effective_category_name_map
    category_name_map = await get_effective_category_name_map(db)

    # Extract unique Sunsky categories
    cat_counts: dict[str, int] = {}
    for p in products:
        raw = p.raw_data or {}
        cat = _extract_sunsky_cat(raw, category_name_map)
        if cat:
            cat_counts[cat] = cat_counts.get(cat, 0) + 1

    # Load saved mappings for this store
    saved_rows = (
        await db.execute(
            select(SunskyCategoryMapping).where(
                SunskyCategoryMapping.store_id == pl.store_id
            )
        )
    ).scalars().all()
    saved: dict[str, SunskyCategoryMapping] = {r.sunsky_cat: r for r in saved_rows}

    # Load WooCommerce categories — include parent_id for tree display
    woo_cats = (
        await db.execute(
            select(WooCategory).where(WooCategory.store_id == pl.store_id)
        )
    ).scalars().all()

    # Load attribute profiles for the panel B dropdown
    from sqlalchemy.orm import selectinload
    profiles = (
        await db.execute(
            select(AttributeProfile)
            .options(selectinload(AttributeProfile.attributes))
            .order_by(AttributeProfile.name)
        )
    ).scalars().all()

    # Total product count in this batch (products may have no extractable category)
    total_products = (await db.execute(
        select(func.count(Product.id)).where(Product.fetch_job_id == pl.fetch_job_id)
    )).scalar_one()

    categories = []
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        m = saved.get(cat)
        woo_cat_list = _mapping_woo_cats(m) if m else []
        primary_id = m.primary_woo_cat_id if m else (woo_cat_list[0]["id"] if woo_cat_list else None)
        categories.append({
            "sunsky_cat":         cat,
            "product_count":      count,
            "woo_cats":           woo_cat_list,
            "primary_woo_cat_id": primary_id,
            "profile_id":         m.profile_id if m else None,
            "is_new":             m is None,
            "times_used":         m.times_used if m else 0,
        })

    # ── CSV import fallback ──────────────────────────────────────────────────
    # When products have no Sunsky category in raw_data (e.g. CSV-sourced
    # pipelines), cat_counts is empty.  Inject a synthetic entry so the user
    # can still assign a WooCommerce category to the whole batch.
    if not categories and total_products > 0:
        fetch_job = await db.get(Job, pl.fetch_job_id) if pl.fetch_job_id else None
        is_csv = fetch_job and fetch_job.type == JobType.csv_import
        label = "CSV Import" if is_csv else "Uncategorised Products"
        m = saved.get(label)
        woo_cat_list = _mapping_woo_cats(m) if m else []
        primary_id = m.primary_woo_cat_id if m else (woo_cat_list[0]["id"] if woo_cat_list else None)
        categories.append({
            "sunsky_cat":         label,
            "product_count":      total_products,
            "woo_cats":           woo_cat_list,
            "primary_woo_cat_id": primary_id,
            "profile_id":         m.profile_id if m else None,
            "is_new":             m is None,
            "times_used":         m.times_used if m else 0,
        })

    return {
        "pipeline_id":    pipeline_id,
        "store_id":       pl.store_id,
        "total_products": total_products,
        "categories":     categories,
        "woo_options": [
            {"id": c.woo_id, "name": c.name, "parent_id": c.parent_id or 0}
            for c in sorted(woo_cats, key=lambda x: x.name)
        ],
        "profiles": [
            {
                "id": p.id,
                "name": p.name,
                "description": p.description,
                "attributes": [
                    {"woo_attr_name": a.woo_attr_name, "required": a.required}
                    for a in (p.attributes or [])
                ],
            }
            for p in profiles
        ],
    }


@router.post("/pipelines/{pipeline_id}/map-confirm")
async def map_confirm(
    pipeline_id: int,
    req: MapConfirmRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Save multi-category mappings and resume the pipeline (Upload → Sync).

    Each entry carries woo_cats (full set) + primary_woo_cat_id.
    Backward-compat columns woo_cat_id / woo_cat_name are updated from the primary.
    Only entries with save_as_rule=True are persisted to the dictionary.
    All entries (saved or not) are applied to this pipeline run.
    """
    pl = await db.get(PipelineJob, pipeline_id)
    if not pl:
        raise HTTPException(status_code=404, detail="Pipeline not found")
    if pl.status != "review":
        raise HTTPException(status_code=400, detail=f"Pipeline is not in review state (current: {pl.status})")

    # Build resolved-name → raw Sunsky category ID, from this pipeline's own
    # products, so each mapping row can be saved with sunsky_cat_id filled in
    # immediately rather than staying NULL until Sync's Step A runs and
    # backfills it after the fact. Without this, Content Review's display
    # (which matches by ID, see pipeline.py's content-data) can't find a
    # mapping that was, in fact, just saved correctly.
    from services.enrich_service import get_effective_category_name_map
    category_name_map = await get_effective_category_name_map(db)
    name_to_cat_id: dict[str, str] = {}
    batch_products = (
        await db.execute(select(Product).where(Product.fetch_job_id == pl.fetch_job_id))
    ).scalars().all()
    for p in batch_products:
        raw = p.raw_data or {}
        resolved_name = _extract_sunsky_cat(raw, category_name_map)
        cat_id = _extract_sunsky_cat_id(raw)
        if resolved_name and cat_id:
            name_to_cat_id.setdefault(resolved_name.strip().lower(), cat_id)

    for entry in req.mappings:
        if not entry.sunsky_cat or not entry.woo_cats:
            continue

        # Resolve primary category
        primary_id = entry.primary_woo_cat_id or (entry.woo_cats[0].id if entry.woo_cats else None)
        primary_cat = next((c for c in entry.woo_cats if c.id == primary_id), entry.woo_cats[0] if entry.woo_cats else None)

        cats_json = json.dumps([{"id": c.id, "name": c.name} for c in entry.woo_cats])
        profile_id = entry.profile_id or None
        sunsky_cat_id = name_to_cat_id.get(entry.sunsky_cat.strip().lower())

        if entry.save_as_rule:
            stmt = (
                pg_insert(SunskyCategoryMapping)
                .values(
                    store_id=pl.store_id,
                    sunsky_cat=entry.sunsky_cat,
                    sunsky_cat_id=sunsky_cat_id,
                    woo_cat_id=primary_cat.id if primary_cat else None,
                    woo_cat_name=primary_cat.name if primary_cat else None,
                    woo_cats_json=cats_json,
                    primary_woo_cat_id=primary_id,
                    profile_id=profile_id,
                    times_used=1,
                    last_used_at=datetime.now(timezone.utc),
                    updated_at=datetime.now(timezone.utc),
                )
                .on_conflict_do_update(
                    index_elements=["store_id", "sunsky_cat"],
                    set_={
                        "sunsky_cat_id":      sunsky_cat_id,
                        "woo_cat_id":         primary_cat.id if primary_cat else None,
                        "woo_cat_name":       primary_cat.name if primary_cat else None,
                        "woo_cats_json":      cats_json,
                        "primary_woo_cat_id": primary_id,
                        "profile_id":         profile_id,
                        "times_used":         SunskyCategoryMapping.__table__.c.times_used + 1,
                        "last_used_at":       datetime.now(timezone.utc),
                        "updated_at":         datetime.now(timezone.utc),
                    },
                )
            )
            await db.execute(stmt)
        else:
            stmt = (
                pg_insert(SunskyCategoryMapping)
                .values(
                    store_id=pl.store_id,
                    sunsky_cat=entry.sunsky_cat,
                    sunsky_cat_id=sunsky_cat_id,
                    woo_cat_id=primary_cat.id if primary_cat else None,
                    woo_cat_name=primary_cat.name if primary_cat else None,
                    woo_cats_json=cats_json,
                    primary_woo_cat_id=primary_id,
                    profile_id=profile_id,
                    times_used=0,
                    last_used_at=datetime.now(timezone.utc),
                    updated_at=datetime.now(timezone.utc),
                )
                .on_conflict_do_update(
                    index_elements=["store_id", "sunsky_cat"],
                    set_={
                        "sunsky_cat_id":      sunsky_cat_id,
                        "woo_cats_json":      cats_json,
                        "primary_woo_cat_id": primary_id,
                        "woo_cat_id":         primary_cat.id if primary_cat else None,
                        "woo_cat_name":       primary_cat.name if primary_cat else None,
                        "profile_id":         profile_id,
                        "updated_at":         datetime.now(timezone.utc),
                    },
                )
            )
            await db.execute(stmt)

    await db.commit()

    # Transition to content_review so the user can review generated content before upload
    pl.status = "content_review"
    pl.updated_at = datetime.now(timezone.utc)
    await db.commit()

    from models.models import PipelineLog
    db.add(PipelineLog(
        pipeline_job_id=pipeline_id, level="info",
        message="Category mapping confirmed — pausing for content review before upload",
    ))
    await db.commit()

    return {"ok": True, "pipeline_id": pipeline_id, "mapped": len(req.mappings)}


# ─────────────────────────────────────────────────────────────────────────────
# Store-scoped endpoints (Settings page — Category mapping dictionary)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/stores/{store_id}/category-mappings")
async def list_category_mappings(store_id: int, db: AsyncSession = Depends(get_db)):
    from models.models import AttributeProfile

    rows = (
        await db.execute(
            select(SunskyCategoryMapping)
            .where(SunskyCategoryMapping.store_id == store_id)
            .order_by(SunskyCategoryMapping.sunsky_cat)
        )
    ).scalars().all()

    profile_ids = {r.profile_id for r in rows if r.profile_id}
    profile_names: dict[int, str] = {}
    if profile_ids:
        profile_rows = (
            await db.execute(select(AttributeProfile).where(AttributeProfile.id.in_(profile_ids)))
        ).scalars().all()
        profile_names = {p.id: p.name for p in profile_rows}

    return {
        "store_id": store_id,
        "mappings": [
            {
                "id":                 r.id,
                "sunsky_cat":         r.sunsky_cat,
                "woo_cats":           _mapping_woo_cats(r),
                "primary_woo_cat_id": r.primary_woo_cat_id or r.woo_cat_id,
                "profile_id":         r.profile_id,
                "profile_name":       profile_names.get(r.profile_id) if r.profile_id else None,
                "times_used":         r.times_used or 0,
                "last_used_at":       r.last_used_at.isoformat() if r.last_used_at else None,
                "updated_at":         r.updated_at.isoformat() if r.updated_at else None,
            }
            for r in rows
        ],
    }


@router.put("/stores/{store_id}/category-mappings")
async def update_category_mappings(
    store_id: int,
    entries: list[CategoryMappingUpdate],
    db: AsyncSession = Depends(get_db),
):
    for entry in entries:
        if not entry.sunsky_cat:
            continue
        primary_id = entry.primary_woo_cat_id or (entry.woo_cats[0].id if entry.woo_cats else None)
        primary_cat = next((c for c in entry.woo_cats if c.id == primary_id), entry.woo_cats[0] if entry.woo_cats else None)
        cats_json = json.dumps([{"id": c.id, "name": c.name} for c in entry.woo_cats])

        stmt = (
            pg_insert(SunskyCategoryMapping)
            .values(
                store_id=store_id,
                sunsky_cat=entry.sunsky_cat,
                woo_cat_id=primary_cat.id if primary_cat else None,
                woo_cat_name=primary_cat.name if primary_cat else None,
                woo_cats_json=cats_json,
                primary_woo_cat_id=primary_id,
                profile_id=entry.profile_id,
                times_used=0,
                updated_at=datetime.now(timezone.utc),
            )
            .on_conflict_do_update(
                index_elements=["store_id", "sunsky_cat"],
                set_={
                    "woo_cat_id":         primary_cat.id if primary_cat else None,
                    "woo_cat_name":       primary_cat.name if primary_cat else None,
                    "woo_cats_json":      cats_json,
                    "primary_woo_cat_id": primary_id,
                    "profile_id":         entry.profile_id,
                    "updated_at":         datetime.now(timezone.utc),
                },
            )
        )
        await db.execute(stmt)
    await db.commit()
    return {"ok": True, "saved": len(entries)}


@router.post("/stores/{store_id}/category-mappings/import")
async def import_category_mappings_file(
    store_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Import category mappings from an Excel (.xlsx) or CSV file.

    Required columns (case-insensitive header matching):
      - "Sunsky Category"  — the Sunsky category ID or name
      - "Woo Category"     — the WooCommerce category name (must already be synced)

    Multiple Woo categories per Sunsky category can be specified by repeating
    the Sunsky Category value on consecutive rows — all rows for the same
    Sunsky category are merged into a single multi-category mapping.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    # ── Parse rows from file ─────────────────────────────────────────────────
    raw_rows: list[tuple[str, str]] = []   # (sunsky_cat, woo_name)
    parse_error: Optional[str] = None

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            sunsky_col = next((i for i, h in enumerate(headers) if "sunsky" in h), None)
            woo_col    = next((i for i, h in enumerate(headers) if "woo" in h), None)
            if sunsky_col is None or woo_col is None:
                parse_error = "Excel must have columns 'Sunsky Category' and 'Woo Category'"
            else:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    s = str(row[sunsky_col] or "").strip()
                    w = str(row[woo_col]    or "").strip()
                    if s and w:
                        raw_rows.append((s, w))
        except Exception as exc:
            parse_error = f"Could not read Excel file: {exc}"
    elif filename.endswith(".csv"):
        try:
            reader = csv.reader(io.StringIO(content.decode("utf-8-sig")))
            headers = [h.strip().lower() for h in next(reader, [])]
            sunsky_col = next((i for i, h in enumerate(headers) if "sunsky" in h), None)
            woo_col    = next((i for i, h in enumerate(headers) if "woo" in h), None)
            if sunsky_col is None or woo_col is None:
                parse_error = "CSV must have columns 'Sunsky Category' and 'Woo Category'"
            else:
                for row in reader:
                    if len(row) > max(sunsky_col, woo_col):
                        s = row[sunsky_col].strip()
                        w = row[woo_col].strip()
                        if s and w:
                            raw_rows.append((s, w))
        except Exception as exc:
            parse_error = f"Could not read CSV file: {exc}"
    else:
        parse_error = "Unsupported file type — upload .xlsx or .csv"

    if parse_error:
        raise HTTPException(400, parse_error)

    if not raw_rows:
        raise HTTPException(400, "No data rows found in the file")

    # ── Load WooCommerce categories for matching ──────────────────────────────
    woo_cats = (
        await db.execute(select(WooCategory).where(WooCategory.store_id == store_id))
    ).scalars().all()
    woo_by_name: dict[str, WooCategory] = {c.name.strip().lower(): c for c in woo_cats}

    # ── Group rows by Sunsky category (support multi-Woo-cat per Sunsky cat) ─
    from collections import OrderedDict
    grouped: dict[str, list[WooCategory]] = OrderedDict()
    skipped: list[str] = []

    for sunsky_cat, woo_name in raw_rows:
        woo_cat = woo_by_name.get(woo_name.strip().lower())
        if not woo_cat:
            skipped.append(f"Row skipped — Woo category '{woo_name}' not found for Sunsky '{sunsky_cat}'")
            continue
        if sunsky_cat not in grouped:
            grouped[sunsky_cat] = []
        # Avoid duplicate Woo cats for the same Sunsky cat
        if not any(c.woo_id == woo_cat.woo_id for c in grouped[sunsky_cat]):
            grouped[sunsky_cat].append(woo_cat)

    # ── Upsert mappings ───────────────────────────────────────────────────────
    imported = 0
    for sunsky_cat, woo_cat_list in grouped.items():
        primary_cat = woo_cat_list[0]
        cats_json = json.dumps([{"id": c.woo_id, "name": c.name} for c in woo_cat_list])
        stmt = (
            pg_insert(SunskyCategoryMapping)
            .values(
                store_id=store_id,
                sunsky_cat=sunsky_cat,
                woo_cat_id=primary_cat.woo_id,
                woo_cat_name=primary_cat.name,
                woo_cats_json=cats_json,
                primary_woo_cat_id=primary_cat.woo_id,
                times_used=0,
                updated_at=datetime.now(timezone.utc),
            )
            .on_conflict_do_update(
                index_elements=["store_id", "sunsky_cat"],
                set_={
                    "woo_cat_id":         primary_cat.woo_id,
                    "woo_cat_name":       primary_cat.name,
                    "woo_cats_json":      cats_json,
                    "primary_woo_cat_id": primary_cat.woo_id,
                    "updated_at":         datetime.now(timezone.utc),
                },
            )
        )
        await db.execute(stmt)
        imported += 1

    await db.commit()
    return {
        "ok":         True,
        "imported":   imported,
        "skipped":    skipped,
        "total_rows": len(raw_rows),
    }


@router.delete("/stores/{store_id}/category-mappings/{mapping_id}")
async def delete_category_mapping(
    store_id: int,
    mapping_id: int,
    db: AsyncSession = Depends(get_db),
):
    row = await db.get(SunskyCategoryMapping, mapping_id)
    if not row or row.store_id != store_id:
        raise HTTPException(status_code=404, detail="Mapping not found")
    await db.delete(row)
    await db.commit()
    return {"ok": True}
